from openpyxl import load_workbook
import pymysql
import os
import datetime

wb = load_workbook("output_data.xlsx")
sheet = wb['Данные']

try:
    # Подключение к БД
    HeidiSQLbd = pymysql.connect(
        host='',
        user='root',
        password='',
        database='tst'
    )
    cursor = HeidiSQLbd.cursor()
    existingvalues = set()
    cursor.execute("SELECT licenc_number FROM keys_lic")
    existingvalues = {row[0] for row in cursor.fetchall()}
    
    def remove_apostrophe(value):
        if value.startswith("'"):
            value = value[1:]
        return value  
    
    unique_values = {}
    
    # Определение уникальных значений и словаря для хранения их id
    unique_values = {}
    existingvalues = {}

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i != 0:
            # Инструкция из какого столбца положить в таблицу
            if row[6] is not None and row[6] != '' and 'нет значения' not in row[6]:
                platform = str(row[10])
                record_link = str(row[1])
                name_countrparty = str(row[3])
                record_names = str(row[7]).lstrip('\'')
                licenc = str(row[6]).lstrip('\'')
                volume_parameters = str(row[11])
                users = str(row[12])
                subscription_upgrade = str(row[9])
                date_creation = str(row[8])

                # Загрузка данных в таблицу keys_lic
                insertquerykeylic = "INSERT INTO keys_lic (licenc_number) VALUES (%s) ON DUPLICATE KEY UPDATE licenc_number=VALUES(licenc_number)"
                cursor.execute(insertquerykeylic, (licenc,))
                key_lic_id = cursor.lastrowid or existingvalues.get(licenc, None) 

                # Загрузка данных в таблицу records_doc
                existingvalues[licenc] = key_lic_id
                formatted_datetime = datetime.datetime.now().strftime('%d.%m.%Y %H:%M')
                insertqueryrecords = "INSERT INTO records_doc (id_key_lic, record_name, record_link, name_countrparty, volume_parameters, users, subscription_upgrade, date_creation, platdorm, date_change, computer_name, user_name) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                cursor.execute(insertqueryrecords, (key_lic_id, record_names, record_link, name_countrparty, volume_parameters, users, subscription_upgrade, date_creation, platform, formatted_datetime, os.getenv('COMPUTERNAME'), os.getenv('USERNAME')))
                # insert = "INSERT INTO function (id_key_lic, volume_parameters, users, subscription_upgrade, date_creation, platdorm) VALUES (%s, %s, %s, %s, %s, %s)"
                # cursor.execute(insert, (key_lic_id, volume_parameters, users, subscription_upgrade, date_creation, platform))
                HeidiSQLbd.commit()
    # Закрытие сеансов
    cursor.close()
    HeidiSQLbd.close()

# Вывод ошибок
except pymysql.Error as error:
    print(error)

except IndexError:
    print("Index Error: Accessing invalid index in tuple")



