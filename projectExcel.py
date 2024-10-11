import os
import openpyxl
import xlrd
import re
import subprocess
from openpyxl import Workbook
from datetime import datetime

# -*- coding: UTF-8 -*-
processed_files_count = 0
root_folder = r"c:\Users\medvedev\Desktop\документы\все документы"
errors = []
error_list = []
no_errors = []

# Функция для проверки наличия значения "Наименование изделия" в столбце B
def check_value_in_columnD(file_path):
    try:
        with open(file_path, 'rb') as file:
            if file_path.endswith(('.xlsx', '.xlsm')):
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=1, max_col=4): 
                    if "Упаковка" in str(row[3].value):
                        return True
                return False                    
            elif file_path.endswith('.xls'):
                wb = xlrd.open_workbook(file_path)
                sheet = wb.sheet_by_index(0)
                for row_idx in range(sheet.nrows):
                    row_data = sheet.row_values(row_idx)
                    if "Упаковка" in str(row_data[3]):
                        
                        return True
                return False
    except Exception as e:
        print(f"Ошибка при проверке файла {file_path}: {e}")
        return False

# Рекурсивная функция для обхода всех папок и файлов
def process_file(file_path, output_sheet):
    global processed_files_count
    if check_value_in_columnD(file_path):
        processed_files_count = 0  
        
        print(f"Содержимое файла {file_path}:")
        mtime = os.path.getmtime(file_path)
        last_modified_date = datetime.fromtimestamp(mtime).strftime('%d.%m.%Y %H:%M:%S')
        open_date = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        try:
            with open(file_path, 'r+') as file:
                
                if file_path.endswith(('.xlsx', '.xlsm')):
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    sheet = wb.active
                    found_start = False
                    processed_files_count += 1
                    print("file xlsx or xlsm open")
                    for row in sheet.iter_rows(values_only=True):
                        if not found_start:
                            if "Наименование изделия" in str(row[1]):
                                for row in row:
                                    if str(row).strip() == "":
                                        continue
                                    elif not str(row).strip() == "":
                                            found_start = True
                                print("True C")
                        else:
                            if not row or not any(row):
                                continue
                            elif "Итого" in str(row[1]):
                                break
                            elif found_start:
                                for the_consignee_idx in range(2, sheet.max_row): 
                                    the_consignee_data = sheet.cell(row=the_consignee_idx, column=1).value
                                    if "Грузополучатель " in str(the_consignee_data):
                                        the_consignee_data = the_consignee_data.split(':')[1] or the_consignee_data.split(': ')[1]
                                        print("True D")
                                        for row_num in range(2, 4):
                                            for col_num in range(10, 13):
                                                cell_value = sheet.cell(row=row_num, column=col_num).value
                                                
                                                if cell_value:
                                                    match = ''
                                                    number_doc = ''
                                                    output_row_idx = len(output_sheet['A']) + 1
                                                    if '\"ОИК' in str(row[1]):
                                                        flag_OIK = 1
                                                        def replace_in_english(text):
                                                            text = text.replace('А', 'A')
                                                            text = text.replace('В', 'B')
                                                            text = text.replace('С', 'C')
                                                            text = text.replace('Е', 'E')
                                                            return text
                                                        print(flag_OIK)
                                                        if re.search(r'\b[A-F0-9АВСЕ]{16}\b', str(row[1])):
                                                            match = re.search(r'\b[A-F0-9АВСЕ]{16}\b', str(row[1])).group()
                                                            match = replace_in_english(match)
                                                            match = '\'' + match
                                                            if re.search(r'\b[A-F0-9(-]{6}\b', str(file_path)):
                                                                number_doc = re.search(r'\b[A-F0-9(-]{6}\b', str(file_path)).group()
                                                                number_doc = '\'' + number_doc
                                                                print(number_doc)
                                                            else:
                                                                number_doc = 'нет значения'
                                                                print(number_doc)

                                                        elif re.search(r'\b[ID№A-F0-9АВСЕ,,+:.]{16}\b', str(row[1])):
                                                            match = re.search(r'\b[ID№A-F0-9АВСЕ,,+:.]{16}\b', str(row[1])).group()
                                                            match = replace_in_english(match)
                                                            match = '\'' + match
                                                            if re.search(r'\b[A-F0-9(-]{6}\b', str(file_path)):
                                                                number_doc = re.search(r'\b[A-F0-9(-]{6}\b', str(file_path)).group()
                                                                number_doc = '\'' + number_doc
                                                                print(number_doc)
                                                            else:
                                                                number_doc = 'нет значения'
                                                                print(number_doc)
                                                        elif re.search(r'\b[A-F0-9(-]{6}\b', str(file_path)):
                                                                number_doc = re.search(r'\b[A-F0-9(-]{6}\b', str(file_path)).group()
                                                                number_doc = '\'' + number_doc
                                                                print(number_doc)
                                                                match = 'нет значения'
                                                                print(match)
                                                            
                                                        
                                                        else:
                                                            match = 'нет значения'
                                                            number_doc = 'нет значения'
                                                            print(number_doc)
                                                            print(match)
                                                            
                                                    else:
                                                        flag_OIK = 0
                                                    cell_value = str(cell_value)    
                                                    if cell_value.endswith('г.'):
                                                        cell_value = cell_value[:-2]
                                                        print(cell_value)
                                                    else:
                                                        print(cell_value)
                                                        continue
                                                    output_sheet.append([output_row_idx - 1, file_path, row[1], the_consignee_data, cell_value, flag_OIK, match, number_doc])
                                                    print("True E")
                                                    break
            close_date = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
            new_sheet.append([file_path, last_modified_date, open_date, close_date, 'Успешно!']) 
            processed_files_count += 1

                    

                    
        except PermissionError as e:
            error_list = [f'{e}']
            close_date = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
            new_sheet.append([file_path, last_modified_date, open_date, close_date, ', '.join(error_list)]) 

        
                
        try:
            with open(file_path, 'r+') as file:
                if os.path.isfile(file_path) and os.access(file_path, os.W_OK):
                    if file_path.endswith('.xls'):
                        wb = xlrd.open_workbook(file_path)
                        sheet = wb.sheet_by_index(0)
                        found_start = False
                        print("file xls open")
                        
                        for row_idx in range(sheet.nrows):
                            row_data = sheet.row_values(row_idx)
                            if not found_start:
                                if "Наименование изделия" in str(row_data[1]):
                                        for row in row_data:
                                            if row.strip() == "":
                                                continue
                                            elif not row.strip() == "":
                                                    found_start = True
                                        print("True C")
                            else:
                                if not row_data or not any(row_data):
                                    continue
                                elif "Итого" in str(row_data[1]):
                                    break


                                elif found_start:
                                    for the_consignee_idx in range(2, sheet.nrows): 
                                        the_consignee_data = sheet.row_values(the_consignee_idx)
                                        if "Грузополучатель" in str(the_consignee_data[0]):
                                            the_consignee_data = the_consignee_data[0].split(':')[1]
                                            print("True D")
                                            for row_num in range(2, 3):
                                                for col_num in range(10, 13):
                                                    if col_num < sheet.ncols and row_num < sheet.nrows: 
                                                        cell_value = sheet.cell_value(row_num, col_num)
                                                        if cell_value:
                                                            match = ''
                                                            number_doc = ''
                                                            output_row_idx = len(output_sheet['A']) + 1
                                                            if '\"ОИК' in str(row_data[1]):
                                                                flag_OIK = 1
                                                                def replace_in_english(text):
                                                                    text = text.replace('А', 'A')
                                                                    text = text.replace('В', 'B')
                                                                    text = text.replace('С', 'C')
                                                                    text = text.replace('Е', 'E')
                                                                    return text
                                                                print(flag_OIK)
                                                                if re.search(r'\b[A-F0-9АВСЕ]{16}\b', str(row_data[1])):
                                                                    match = re.search(r'\b[A-F0-9АВСЕ]{16}\b', str(row_data[1])).group()
                                                                    match = replace_in_english(match)
                                                                    match = '\'' + match
                                                                    if re.search(r'\b[A-F0-9(-]{6}\b', str(file_path)):
                                                                        number_doc = re.search(r'\b[A-F0-9(-]{6}\b', str(file_path)).group()
                                                                        number_doc = '\'' + number_doc
                                                                        print(number_doc)
                                                                    else:
                                                                        number_doc = 'нет значения'
                                                                        print(number_doc)
                                                                elif re.search(r'\b[ID№A-F0-9АВСЕ,,+:.]{16}\b', str(row_data[1])):
                                                                    match = re.search(r'\b[ID№A-F0-9АВСЕ,,+:.]{16}\b', str(row_data[1])).group()
                                                                    match = replace_in_english(match)
                                                                    match = '\'' + match
                                                                    if re.search(r'\b[A-F0-9(-]{6}\b', str(file_path)):
                                                                        number_doc = re.search(r'\b[A-F0-9(-]{6}\b', str(file_path)).group()
                                                                        number_doc = '\'' + number_doc
                                                                        print(number_doc)
                                                                    else:
                                                                        number_doc = 'нет значения'
                                                                        print(number_doc)    
                                                                elif re.search(r'\b[A-F0-9(-]{6}\b', str(file_path)):
                                                                        number_doc = re.search(r'\b[A-F0-9(-]{6}\b', str(file_path)).group()
                                                                        number_doc = '\'' + number_doc
                                                                        print(number_doc)
                                                                        match = 'нет значения'
                                                                        print(match)
                                                                    
                                                                
                                                                else:
                                                                    match = 'нет значения'
                                                                    number_doc = 'нет значения'
                                                                    print(number_doc)
                                                                    print(match)
                                                                    
                                                            else:
                                                                flag_OIK = 0
                                                            cell_value = str(cell_value)    
                                                            if cell_value.endswith('г.'):
                                                                cell_value = cell_value[:-2]
                                                                print(cell_value)
                                                            else:
                                                                print(cell_value)
                                                                continue
                                                            output_sheet.append([(output_row_idx - 1), file_path, row_data[1], the_consignee_data, cell_value, flag_OIK, match, number_doc])  
                                                            

                                                            print("True E")
                                                            break
            close_date = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
            new_sheet.append([file_path, last_modified_date, open_date, close_date, 'Успешно!']) 
            processed_files_count += 1        
                    
        except PermissionError as e:
            print('False')
            error_list = [f'{e}']
            close_date = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
            new_sheet.append([file_path, last_modified_date, open_date, close_date, ', '.join(error_list)]) 
                                               
            



# Создаем новый Excel файл
output_wb = Workbook()
output_sheet = output_wb.active

output_sheet.title = "Данные"
output_sheet.append(["№", "Путь к файлу", "Наименование изделия", "Грузополучатель", "Дата из файла", "Признак ОИК", "Лицензия", "Номер договора"])

new_sheet = output_wb.create_sheet(title="Перечень файлов")
new_sheet.append(["Полный путь", "Дата изменения файла", "Дата начала обработки", "Дата окончания обработки", "Комментарий"])

sheet_names = output_wb.sheetnames
output_wb._sheets = [output_wb[sheet_name] for sheet_name in reversed(sheet_names)]


# Рекурсивная функция для обхода всех папок и файлов
def process_files_in_folder(folder, output_sheet):
    for root, dirs, files in os.walk(folder):
        for filename in files:
            if filename.endswith(('.xls', '.xlsx', '.xlsm')):
                file_path = os.path.join(root, filename)
                process_file(file_path, output_sheet)

# Вызываем функцию для обработки всех файлов в указанной папке и подпапках
process_files_in_folder(root_folder, output_sheet)
subprocess.call(["python", "project_pars2_excel.py"])
# Сохраняем созданный Excel файл
output_wb.save("output_data.xlsx")
output_wb.close()

print(f"Files {processed_files_count} processed.")