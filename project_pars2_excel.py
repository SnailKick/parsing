import xlrd
import openpyxl

doc_1 = xlrd.open_workbook("Отчет 27.05.2024.xls")
doc_2 = openpyxl.load_workbook("output_data.xlsx")
sheet1 = doc_1.sheet_by_index(0)
sheet2 = doc_2['Данные']
headers = ['Дата создания', 'Подписка и обн.', 'Платформа', 'Объем парам.', 'Пользователи']
for i, header in enumerate(headers, start=9):
    sheet2.cell(row=1, column=i, value=header)

try:
    col_a = [cell.value.strip().replace(";", "").replace(",", "").replace(" ", "") for cell in sheet1.col(0)]
    col_g = [str(cell.value).strip().replace("'", "").replace(",", "").replace(" ", "") for row in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, min_col=7, max_col=7) for cell in row]
    processed_files_count = 0
    for val_a in col_a:
        for val_g in col_g:
            if val_a == val_g:
                for i, val_a in enumerate(col_a, start=1):
                    for j, val_g in enumerate(col_g, start=2):
                        if val_a == val_g:
                            row_num = j
                            # Извлечение данных из doc_1
                            data_F = str(sheet1.cell_value(i, 5))
                            data_G = str(sheet1.cell_value(i, 6))
                            data_I = str(sheet1.cell_value(i, 8))
                            data_K = str(sheet1.cell_value(i, 10))
                            data_M = str(sheet1.cell_value(i, 12))

                            # Запись данных в doc_2
                            sheet2.cell(row=row_num, column=9, value=data_F)
                            sheet2.cell(row=row_num, column=10, value=data_G)
                            sheet2.cell(row=row_num, column=11, value=data_I)
                            sheet2.cell(row=row_num, column=12, value=data_K)
                            sheet2.cell(row=row_num, column=13, value=data_M)
                            processed_files_count += 1
except Exception as error:
    print(error)

print(processed_files_count)
print(f"{processed_files_count} rows processed successfully.")
doc_1.release_resources()